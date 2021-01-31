import csv
import openpyxl
from openpyxl.utils import get_column_letter

'''
Data Example
CSV Data
Component Name, BOM
Microchip 1, NI
Resistor 2, I
Capacitor 3, NI
'''

def get_list_of_dict_from_rpt(file_path):
    with open(file_path, newline='') as file:
        # Get list of field names
        for i in range(6):
            fields = file.readline()
            if i == 5:
                field_names = fields[:-1].split("#")
        
        reader = csv.DictReader(file, fieldnames=field_names, delimiter='#')

        rpt_list = []
        for row in reader:
            rpt_list.append(row)
    return rpt_list

def get_installed_components_from_dict_list(rpt_list):
    installed = []
    for item in rpt_list:
        if item['BOM'] == 'I' or item['BOM'] == 'DEBUG':
            installed.append(item)
    return installed

def get_not_installed_components_from_dict_list(rpt_list):
    not_installed = []
    for item in rpt_list:
        if item['BOM'] == 'NI':
            not_installed.append(item)
    return not_installed

def get_everything_else_from_dict_list(rpt_list):
    everything_else = []
    for item in rpt_list:
        #if item['BOM'] != 'NI' and item['BOM'] != "I" and item['BOM'] != "DEBUG":
        if item['BOM'] != 'NI' and item['BOM'] != "I" :
            everything_else.append(item)
    return everything_else

def write_to_csv(dict_list, output_path):
    keys = dict_list[0].keys()
    with open(output_path, 'w', newline='') as file:
        dict_writer = csv.DictWriter(file, keys)
        dict_writer.writeheader()
        dict_writer.writerows(dict_list)

def write_to_ws(dict_list, ws_path):
    col_letter_list = list( upper_letter_range("A","P") )
    #col_letter_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']
    if len(dict_list) == 0:
        return 0

    keys = dict_list[0].keys()
    for index,col_letter in enumerate(col_letter_list):
        s_index = f"{col_letter}{1}"  # row 1 is the header
        ws_path[s_index] = list(keys)[index]

    for i in range (len(dict_list) ):
        values = dict_list[i].values()   
        for index,col_letter in enumerate(col_letter_list):
            s_index = f"{col_letter}{i+2}"  # row >=2 is the data
            #ws_path[s_index] = list(values)[index]
            if col_letter == 'A':   # item
                ws_path[s_index] = i+1
            elif col_letter == 'F':  # Quantity
                ws_path[s_index] = int( list(values)[index] )
                ws_path.cell(i+2,6).number_format = '0'
            elif col_letter == 'H':   # CCL
                ws_path[s_index] = int ( list(values)[index] )
                ws_path.cell(i+2,8).number_format = '0'
            else:
                ws_path[s_index] = list(values)[index]

    #format the column widths
    column_widths = []
    for row in ws_path.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        if column_width > 50:
            column_width = 50
        elif column_width < 5:
            column_width = 6

        if i == 1:               # HH_PN
            column_width = 18
        elif i == 2:               # DESCRIPTION
            column_width = 80
        ws_path.column_dimensions[get_column_letter(i + 1)].width = column_width

def as_text(value):
    if value is None:
        return ""
    return str(value)

def upperCaseAlphabets():
    print("Upper Case Alphabets")
    for i in range(65, 91):
        print(chr(i), end=" ")
    print() 

    # another way using format.
    #for i in range(65, 91):   # ASCII number: 97-122 to 'a..z' , 65-91 to 'a..z'
    #  print("{:c}".format(i), end='')
    #print()


def lower_letter_range(start, stop="{", step=1):
    """Yield a range of lowercase letters.""" 
    for ord_ in range(ord(start.lower()), ord(stop.lower()), step):
        yield chr(ord_)   

def upper_letter_range(start, stop="{", step=1):
    """Yield a range of upper letters.""" 
    for ord_ in range(ord(start.upper()), ord(stop.upper()), step):
        yield chr(ord_)   


if __name__ == "__main__":
    from openpyxl import Workbook
    wb = Workbook()

    # grab the active worksheet
    ws1 = wb.create_sheet("ALL",0)
    ws2 = wb.create_sheet("NI",1)
    ws3 = wb.create_sheet("TBD",2)

    # Get data into list of dictionaries where each row is a dictionary
    csvpath = './data/CONAN_MB_EVT1_BOM_20210105.rpt'
    rpt_list = get_list_of_dict_from_rpt(csvpath)
    print(f"The length of the data is: {len(rpt_list)}")

    # Get CSV of installed BOM
    installed_list = get_installed_components_from_dict_list(rpt_list)
    print(f"The length of the installed components are: {len(installed_list)}")
    #installed_output = './installed_BOM.csv'
    #write_to_csv(installed_list, installed_output)
    write_to_ws(installed_list, ws1)

    # Get CSV for non-installed BOM
    not_installed_list = get_not_installed_components_from_dict_list(rpt_list)
    print(f"The length of the non-installed components are: {len(not_installed_list)}")
    #not_installed_output = './not_installed_BOM.csv'
    #write_to_csv(not_installed_list, not_installed_output)
    write_to_ws(not_installed_list, ws2)

    # Get CSV for the rest
    everything_else_list = get_everything_else_from_dict_list(rpt_list)
    print(f"The length of everything else components are: {len(everything_else_list)}")
    #everything_else_output = './everything_else_BOM.csv'
    #write_to_csv(rpt_list, everything_else_output)
    write_to_ws(everything_else_list, ws3)

    wb.save('BOM_output.xlsx')

    # There are BOM equal to DEBUG.
    # for item in rpt_list:
    #     if item['BOM'] != "I" and item['BOM'] != 'NI':
    #         print(item['BOM'])


